#!/usr/bin/env python3
"""
az-summary.py  —  Cortex Cloud Azure Sizing Summary
=====================================================
Reads azure_results.json (produced by az-sizing.py --resume) plus optional
azure_state.jsonl and azure_tenant.json to produce a three-section report
and an Excel workbook with filterable tables.

Usage:
  python3 az-summary.py
  python3 az-summary.py --results azure_results.json
  python3 az-summary.py --results run1.json run2.json   # merge multiple runs
  python3 az-summary.py --state   azure_state.jsonl     # show scan status
  python3 az-summary.py --tenant  azure_tenant.json     # include SaaS users
  python3 az-summary.py --xlsx    sizing_output.xlsx    # export to Excel
  python3 az-summary.py --failed-only --state azure_state.jsonl
"""

import json
import math
import os
import argparse
from datetime import datetime, timezone

SEP  = "═" * 180
THIN = "─" * 180

CC_METERING = {
    "vm": 1, "caas": 10, "serverless": 25,
    "buckets": 10, "db": 2, "images": 10, "saas_users": 10,
}

NAME_W = 38
ID_W   = 36


# ──────────────────────────────────────────────────────────────────────────────
# SKU calculation
# ──────────────────────────────────────────────────────────────────────────────
def _wl(raw: int, rate: int) -> int:
    return math.ceil(raw / rate) if raw > 0 else 0


def compute_skus(raw: dict) -> dict:
    vm_total   = raw.get("vm_running", 0) + raw.get("aks_nodes", 0) + raw.get("aro_nodes", 0)
    caas_total = raw.get("aci_containers", 0) + raw.get("container_app_containers", 0)
    db_total   = raw.get("azure_sql_dbs", 0) + raw.get("cosmos_db", 0)
    # acr_images in JSON = tag count (primary billable unit per official methodology)
    # fallback to acr_images for older result files
    acr_gross  = raw.get("acr_tag_count", raw.get("acr_images", 0))
    net_images = max(0, acr_gross - raw.get("free_image_allowance", 0))

    wl_vm      = _wl(vm_total,                       CC_METERING["vm"])
    wl_caas    = _wl(caas_total,                     CC_METERING["caas"])
    wl_funcs   = _wl(raw.get("azure_functions", 0),  CC_METERING["serverless"])
    wl_db      = _wl(db_total,                       CC_METERING["db"])
    wl_storage = _wl(raw.get("storage_accounts", 0), CC_METERING["buckets"])
    wl_images  = _wl(net_images,                     CC_METERING["images"])
    total      = wl_vm + wl_caas + wl_funcs + wl_db + wl_storage + wl_images

    return {
        "vm_total": vm_total, "caas_total": caas_total, "db_total": db_total,
        "net_images": net_images,
        "wl_vm": wl_vm, "wl_caas": wl_caas, "wl_funcs": wl_funcs,
        "wl_db": wl_db, "wl_storage": wl_storage, "wl_images": wl_images,
        "total": total,
    }


# ──────────────────────────────────────────────────────────────────────────────
# File loaders
# ──────────────────────────────────────────────────────────────────────────────
def load_results(paths: list) -> dict:
    merged = {}
    for path in paths:
        if not os.path.exists(path):
            print(f"  Warning: not found — {path}")
            continue
        with open(path) as fh:
            try:
                data = json.load(fh) or {}
            except json.JSONDecodeError as exc:
                print(f"  Warning: cannot parse {path}: {exc}")
                continue
        merged.update(data)
        print(f"  Loaded {len(data)} subscription(s) from {path}")
    return merged


def load_state(path: str) -> dict:
    state = {}
    if not path or not os.path.exists(path):
        return state
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if line:
                row = json.loads(line)
                state[row["sub_id"]] = row
    return state


def load_tenant(path: str) -> dict:
    if not path or not os.path.exists(path):
        return {}
    with open(path) as fh:
        return json.load(fh) or {}


# ──────────────────────────────────────────────────────────────────────────────
# Formatting helpers
# ──────────────────────────────────────────────────────────────────────────────
def _trunc(s: str, w: int) -> str:
    return s if len(s) <= w else s[:w - 1] + "…"


def _fmt_gb(val) -> str:
    """None → 'NOT AVAILABLE', float → '1.2340 GB/day'."""
    if val is None:
        return "NOT AVAILABLE"
    return f"{val:.4f}"


def _fmt_add(val, included) -> str:
    """Additional ingestion field."""
    if val is None:
        return "n/a"
    add = round(max(0.0, val - included), 4)
    return f"0.0000 ✓" if add == 0.0 else f"{add:.4f} ⚠"


def _pct(part, total) -> str:
    if not total:
        return "—"
    return f"{100 * part / total:.0f}%"


# ──────────────────────────────────────────────────────────────────────────────
# Build flat row list (shared by console + Excel)
# ──────────────────────────────────────────────────────────────────────────────
def build_rows(sorted_items: list, state: dict) -> list:
    """Return a list of dicts, one per subscription, with all fields needed."""
    rows = []
    for sub_id, payload in sorted_items:
        raw    = payload.get("raw_counts", {})
        sku    = compute_skus(raw)
        status = state.get(sub_id, {}).get("status", "") if state else ""
        incl   = round(sku["total"] / 50, 4) if sku["total"] > 0 else 0.0

        audit_gb = raw.get("audit_log_gb_day")
        flow_gb  = raw.get("flow_log_gb_day")

        # Runtime subtotal: only when both audit and flow are measured
        if audit_gb is not None and flow_gb is not None:
            runtime_gb = round(audit_gb + flow_gb, 4)
        else:
            runtime_gb = None

        posture_add = round(max(0.0, audit_gb - incl), 4) if audit_gb is not None else None
        runtime_add = round(max(0.0, runtime_gb - incl), 4) if runtime_gb is not None else None

        rows.append({
            # Identity
            "sub_name":          payload.get("name", sub_id),
            "sub_id":            sub_id,
            "scan_status":       status,
            "scanned_at":        payload.get("timestamp", ""),
            # Raw resource counts
            "vms_running":       raw.get("vm_running", 0),
            "aks_nodes":         raw.get("aks_nodes", 0),
            "aro_nodes":         raw.get("aro_nodes", 0),
            "aci_containers":    raw.get("aci_containers", 0),
            "container_apps":    raw.get("container_app_containers", 0),
            "azure_functions":   raw.get("azure_functions", 0),
            "sql_dbs":           raw.get("azure_sql_dbs", 0),
            "cosmos_dbs":        raw.get("cosmos_db", 0),
            "storage_accounts":  raw.get("storage_accounts", 0),
            "acr_images_gross":  raw.get("acr_tag_count", raw.get("acr_images", 0)),  # tags — primary billable
            "acr_manifest_count": raw.get("acr_manifest_count", 0),                   # unique digests — informational
            "acr_tag_count":     raw.get("acr_tag_count", raw.get("acr_images", 0)),
            "acr_free_allow":    raw.get("free_image_allowance", 0),
            "acr_images_net":    sku["net_images"],
            # Log sources
            "flow_log_sources":  raw.get("flow_log_sources", 0),
            "eh_namespaces":     raw.get("eh_namespaces_found", 0),
            "eh_with_capture":   raw.get("eh_with_capture_found", 0),
            # Workload SKUs
            "wl_vm_nodes":       sku["wl_vm"],
            "wl_caas":           sku["wl_caas"],
            "wl_functions":      sku["wl_funcs"],
            "wl_databases":      sku["wl_db"],
            "wl_storage":        sku["wl_storage"],
            "wl_images":         sku["wl_images"],
            "wl_total":          sku["total"],
            # Log ingestion
            "incl_gb_day":       incl,
            "audit_gb_day":      audit_gb,
            "flow_gb_day":       flow_gb,
            "runtime_gb_day":    runtime_gb,
            "posture_add_gb":    posture_add,
            "runtime_add_gb":    runtime_add,
            "audit_method":      raw.get("audit_log_method", ""),
            "flow_method":       raw.get("flow_log_method", ""),
            # Structured diagnostic records from az-sizing.py (commit 1+)
            "diagnostics":       payload.get("diagnostics") or [],
            # Per-resource inventory of measured log sources (populated by az-sizing.py)
            "flow_log_inventory":    raw.get("flow_log_inventory") or [],
            "eh_capture_inventory":  raw.get("eh_capture_inventory") or [],
        })
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 1 — Workload counts  (console)
# ──────────────────────────────────────────────────────────────────────────────
def print_section1(rows: list) -> dict:
    print(f"\n{SEP}")
    print("  SECTION 1  —  Workload Counts per Subscription")
    print(f"  All resources counted directly via Azure APIs — no estimates")
    print(SEP)

    # Header
    print(
        f"  {'Subscription':<{NAME_W}}  {'Sub ID':<{ID_W}}  "
        f"{'VMs':>5}  {'AKS':>5}  {'ARO':>5}  {'ACI':>5}  {'CApps':>6}  "
        f"{'Funcs':>6}  {'SQL':>5}  {'Cosmos':>7}  {'Store':>6}  "
        f"{'ACR-Tags':>9}  {'ACR-Mfst':>9}  {'ACR-Net':>8}  "
        f"{'WL-VM':>6}  {'WL-CaaS':>8}  {'WL-Func':>8}  "
        f"{'WL-DB':>6}  {'WL-Stor':>8}  {'WL-Img':>7}  {'TOTAL WL':>9}"
    )
    print(THIN)

    # Totals accumulators
    T = {k: 0 for k in [
        "vms_running","aks_nodes","aro_nodes","aci_containers","container_apps",
        "azure_functions","sql_dbs","cosmos_dbs","storage_accounts",
        "acr_tag_count","acr_manifest_count","acr_images_net",
        "wl_vm_nodes","wl_caas","wl_functions","wl_databases","wl_storage","wl_images","wl_total"
    ]}

    for r in rows:
        print(
            f"  {_trunc(r['sub_name'], NAME_W):<{NAME_W}}  {r['sub_id']:<{ID_W}}  "
            f"{r['vms_running']:>5}  {r['aks_nodes']:>5}  {r['aro_nodes']:>5}  "
            f"{r['aci_containers']:>5}  {r['container_apps']:>6}  "
            f"{r['azure_functions']:>6}  {r['sql_dbs']:>5}  {r['cosmos_dbs']:>7}  "
            f"{r['storage_accounts']:>6}  {r['acr_tag_count']:>9}  "
            f"{r['acr_manifest_count']:>9}  {r['acr_images_net']:>8}  "
            f"{r['wl_vm_nodes']:>6}  {r['wl_caas']:>8}  {r['wl_functions']:>8}  "
            f"{r['wl_databases']:>6}  {r['wl_storage']:>8}  {r['wl_images']:>7}  "
            f"{r['wl_total']:>9}"
        )
        for k in T:
            T[k] += r[k]

    print(THIN)
    print(
        f"  {'TOTAL':<{NAME_W}}  {'':<{ID_W}}  "
        f"{T['vms_running']:>5}  {T['aks_nodes']:>5}  {T['aro_nodes']:>5}  "
        f"{T['aci_containers']:>5}  {T['container_apps']:>6}  "
        f"{T['azure_functions']:>6}  {T['sql_dbs']:>5}  {T['cosmos_dbs']:>7}  "
        f"{T['storage_accounts']:>6}  {T['acr_tag_count']:>9}  "
        f"{T['acr_manifest_count']:>9}  {T['acr_images_net']:>8}  "
        f"{T['wl_vm_nodes']:>6}  {T['wl_caas']:>8}  {T['wl_functions']:>8}  "
        f"{T['wl_databases']:>6}  {T['wl_storage']:>8}  {T['wl_images']:>7}  "
        f"{T['wl_total']:>9}"
    )
    print(SEP)

    print(f"\n  Metering ratios:  VM/Node=1  |  CaaS=10 ctrs→1 WL  |  "
          f"Functions=25→1 WL  |  DB=2→1 WL  |  Storage=10→1 WL  |  Images=10→1 WL")
    print(f"  ACR-Tags = total tags (primary billable input — assumes 'All' scan mode)")
    print(f"  ACR-Mfst = unique manifests/digests (informational — deduplication view)")
    print(f"  ACR-Net  = tags after free allowance → drives WL-Img workload count")
    print(f"  ⓘ  If customer uses 'Latest tag' or 'Days Modified' scan mode, ACR workload count will be lower.")

    return T


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 2 — Log ingestion  (console)
# ──────────────────────────────────────────────────────────────────────────────
def print_section2(rows: list) -> dict:
    print(f"\n{SEP}")
    print("  SECTION 2  —  Log Ingestion  (exact blob measurements — no estimates)")
    print(f"  Audit = EH capture blobs (primary) or LAW proxy  |  Flow = VNet/NSG flow-log blobs  (prefix-based, 7-day avg)")
    print(f"  EH is the primary Cortex ingestion path. LAW shown as reference only when both measured.")
    print(f"  Included GB/day = Total WL / 50  |  Flow Logs counted in Runtime SKU only")
    print(SEP)

    print(
        f"  {'Subscription':<{NAME_W}}  {'Sub ID':<{ID_W}}  "
        f"{'EH-NS':>5}  {'EH-Cap':>6}  {'Flow-Src':>9}  "
        f"{'Incl GB/d':>10}  {'Audit GB/d':>11}  {'Flow GB/d':>10}  "
        f"{'Runtime GB/d':>13}  {'Posture+':>10}  {'Runtime+':>10}"
    )
    print(THIN)

    tot_audit = 0.0
    tot_flow  = 0.0
    tot_incl  = 0.0
    miss_audit = miss_flow = False

    for r in rows:
        tot_incl += r["incl_gb_day"]

        audit_str   = _fmt_gb(r["audit_gb_day"])
        flow_str    = _fmt_gb(r["flow_gb_day"])
        runtime_str = _fmt_gb(r["runtime_gb_day"])
        p_add       = _fmt_add(r["audit_gb_day"],   r["incl_gb_day"])
        rt_add      = _fmt_add(r["runtime_gb_day"], r["incl_gb_day"])

        print(
            f"  {_trunc(r['sub_name'], NAME_W):<{NAME_W}}  {r['sub_id']:<{ID_W}}  "
            f"{r['eh_namespaces']:>5}  {r['eh_with_capture']:>6}  {r['flow_log_sources']:>9}  "
            f"{r['incl_gb_day']:>10.4f}  {audit_str:>11}  {flow_str:>10}  "
            f"{runtime_str:>13}  {p_add:>10}  {rt_add:>10}"
        )

        if r["audit_gb_day"] is not None:
            tot_audit += r["audit_gb_day"]
        else:
            miss_audit = True
        if r["flow_gb_day"] is not None:
            tot_flow += r["flow_gb_day"]
        else:
            miss_flow = True

    print(THIN)
    tot_audit_r = round(tot_audit, 4)
    tot_flow_r  = round(tot_flow,  4)
    tot_incl_r  = round(tot_incl,  4)

    # Distinguish "fully unmeasured" (zero subs contributed a value) from
    # "partial" (at least one sub contributed but others did not).
    n_audit_measured = sum(1 for r in rows if r["audit_gb_day"] is not None)
    n_flow_measured  = sum(1 for r in rows if r["flow_gb_day"]  is not None)
    audit_total_str  = (
        "NOT AVAILABLE" if n_audit_measured == 0 else
        f"{tot_audit_r:.4f}  ⚠ partial" if miss_audit else
        f"{tot_audit_r:.4f}"
    )
    flow_total_str   = (
        "NOT AVAILABLE" if n_flow_measured == 0 else
        f"{tot_flow_r:.4f}  ⚠ partial" if miss_flow else
        f"{tot_flow_r:.4f}"
    )

    print(
        f"  {'TOTAL (measured subs only)':<{NAME_W}}  {'':<{ID_W}}  "
        f"{'':>5}  {'':>6}  {'':>9}  "
        f"{tot_incl_r:>10.4f}  {audit_total_str:>11}  {flow_total_str:>10}"
    )
    print(SEP)

    if miss_audit:
        print("  ⚠  Audit logs: NOT AVAILABLE on one or more subscriptions.")
        print("     Action: Enable Event Hub capture + grant Storage Blob Data Reader on capture storage,")
        print("     OR route Activity Logs to a Log Analytics Workspace + grant Log Analytics Reader.")
    if miss_flow:
        print("  ⚠  Flow logs: NOT AVAILABLE on one or more subscriptions.")
        print("     Action: Grant Storage Blob Data Reader on flow-log storage accounts.")
        print("     If role is already assigned, the storage account network firewall may be blocking")
        print("     this client IP — whitelist it in the storage account Networking settings,")
        print("     or run the script from inside the customer's VNet.")

    return {
        "tot_audit":             tot_audit_r,
        "tot_flow":              tot_flow_r,
        "tot_incl":              tot_incl_r,
        "miss_audit":            miss_audit,
        "miss_flow":             miss_flow,
        "fully_unmeasured_audit": n_audit_measured == 0,
        "fully_unmeasured_flow":  n_flow_measured  == 0,
    }


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 3 — Grand total  (console)
# ──────────────────────────────────────────────────────────────────────────────
def print_section3(totals: dict, ingestion: dict, tenant: dict) -> None:
    sub_total  = totals["wl_total"]
    saas_wl    = tenant.get("saas_user_workloads", 0)
    grand_total = sub_total + saas_wl
    grand_incl  = round(grand_total / 50, 4) if grand_total > 0 else 0.0

    print(f"\n{SEP}")
    print("  SECTION 3  —  Grand Total")
    print(SEP)

    print(f"\n  ── Subscription Workloads ───────────────────────────────────────────")
    print(f"  {'VMs / Nodes  (VMs + AKS + ARO)':<48}: {totals['wl_vm_nodes']:>6}")
    print(f"  {'CaaS  (ACI + Container Apps ÷ 10)':<48}: {totals['wl_caas']:>6}")
    print(f"  {'Functions  (Azure Functions ÷ 25)':<48}: {totals['wl_functions']:>6}")
    print(f"  {'Databases  (SQL + Cosmos ÷ 2)':<48}: {totals['wl_databases']:>6}")
    print(f"  {'Storage  (Storage Accounts ÷ 10)':<48}: {totals['wl_storage']:>6}")
    print(f"  {'Images  (ACR net ÷ 10)':<48}: {totals['wl_images']:>6}")
    print(f"  {'─'*56}")
    print(f"  {'Subscription Total':<48}: {sub_total:>6}")

    print(f"\n  ── SaaS Users  (Entra ID — Tenant Level) ───────────────────────────")
    if tenant:
        print(f"  {'Tenant ID':<48}: {tenant.get('tenant_id','—')}")
        print(f"  {'Scanned at':<48}: {tenant.get('timestamp','—')}")
        print(f"  {'Member users (enabled)':<48}: {tenant.get('entra_member_users',0):>6,}")
        print(f"  {'Guest users  (enabled)':<48}: {tenant.get('entra_guest_users',0):>6,}")
        print(f"  {'Total SaaS users':<48}: {tenant.get('total_saas_users',0):>6,}")
        print(f"  {'SaaS Workloads  (÷ 10)':<48}: {saas_wl:>6}")
    else:
        print(f"  Not scanned.  Run: python3 az-sizing.py --tenant-scan")

    print(f"\n  {'═'*56}")
    print(f"  {'GRAND TOTAL Workload Licenses':<48}: {grand_total:>6}")
    print(f"  {'Included log ingestion  (÷ 50)':<48}: {grand_incl:.4f} GB/day")

    # Log sizing
    print(f"\n  ── Log Ingestion Sizing ─────────────────────────────────────────────")
    tot_audit   = ingestion["tot_audit"]
    tot_flow    = ingestion["tot_flow"]
    tot_incl    = ingestion["tot_incl"]
    miss_audit  = ingestion["miss_audit"]
    miss_flow   = ingestion["miss_flow"]
    fully_unm_audit = ingestion.get("fully_unmeasured_audit", miss_audit and tot_audit == 0.0)
    fully_unm_flow  = ingestion.get("fully_unmeasured_flow",  miss_flow  and tot_flow  == 0.0)

    def _avail(v, fully_unmeasured, missing, label):
        if fully_unmeasured:
            return "NOT AVAILABLE"
        s = f"{v:.4f} GB/day"
        if missing:
            s += f"  ⚠ partial ({label} missing on some subs)"
        return s

    print(f"  {'Audit Logs':<48}: {_avail(tot_audit, fully_unm_audit, miss_audit, 'Event Hub capture')}")
    print(f"  {'  └─ Applies to: Posture SKU + Runtime SKU (same volume, not additive)':<48}")
    print(f"  {'Flow Logs  [Runtime SKU only]':<48}: {_avail(tot_flow, fully_unm_flow, miss_flow, 'Blob Reader / flow config')}")
    print(f"  {'─'*56}")

    # Posture — audit only
    if not miss_audit:
        p_add = round(max(0.0, tot_audit - grand_incl), 4)
        p_str = f"{p_add:.4f} GB/day additional" if p_add > 0 else "✓ within included allowance"
        p_meas = f"{tot_audit:.4f}"
    elif fully_unm_audit:
        p_meas = "NOT AVAILABLE"
        p_str  = "cannot determine — audit log volume not measured"
    else:
        p_meas = f"{tot_audit:.4f}"
        p_str  = "cannot determine — audit log volume not fully measured"
    print(f"\n  Posture SKU  —  Audit Logs only:")
    print(f"    Included {grand_incl:.4f} GB/day  |  Measured {p_meas} GB/day  →  {p_str}")

    # Runtime — audit + flow (superset of Posture)
    if not miss_audit and not miss_flow:
        rt_total = round(tot_audit + tot_flow, 4)
        r_add    = round(max(0.0, rt_total - grand_incl), 4)
        r_str    = f"{r_add:.4f} GB/day additional" if r_add > 0 else "✓ within included allowance"
        r_meas   = f"{rt_total:.4f}"
    elif fully_unm_audit and fully_unm_flow:
        r_meas = "NOT AVAILABLE"
        r_str  = "cannot determine — neither audit nor flow logs measured"
    else:
        r_meas = "partially measured"
        r_str  = "cannot determine — audit and/or flow not fully measured"
    print(f"\n  Runtime SKU  —  Audit + Flow Logs  (superset of Posture, not additive):")
    print(f"    Included {grand_incl:.4f} GB/day  |  Measured {r_meas} GB/day  →  {r_str}")
    print(f"  ⓘ  Runtime includes everything Posture measures plus Flow Logs.")
    print(f"     A customer buying Runtime does not also need Posture separately.")
    print(SEP)


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────
def export_excel(rows: list, tenant: dict, ingestion: dict, grand_totals: dict,
                 out_path: str, generated_at: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

    wb = Workbook()

    # ── Color palette ────────────────────────────────────────────────────────
    C_HEADER_BG  = "1F3864"   # Dark navy
    C_HEADER_FG  = "FFFFFF"   # White
    C_SECTION_BG = "2E75B6"   # Mid blue
    C_SECTION_FG = "FFFFFF"
    C_SUBHDR_BG  = "D6E4F0"   # Light blue
    C_SUBHDR_FG  = "1F3864"
    C_TOTAL_BG   = "FFF2CC"   # Pale amber for totals
    C_TOTAL_FG   = "1F3864"
    C_WARN_BG    = "FFE0E0"   # Pale red for missing/warning
    C_GOOD_BG    = "E2EFDA"   # Pale green for good
    C_ALT_BG     = "F5F9FF"   # Alternating row
    C_BORDER     = "BDD7EE"

    thin_side = Side(style="thin", color=C_BORDER)
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    def _hdr_font(bold=True, color=C_HEADER_FG, size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)

    def _right():
        return Alignment(horizontal="right", vertical="center")

    def _set_col_width(ws, col_idx, width):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _style_header_row(ws, row_num, n_cols,
                          bg=C_HEADER_BG, fg=C_HEADER_FG, height=20):
        ws.row_dimensions[row_num].height = height
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=True, color=fg, size=10)
            cell.fill      = _fill(bg)
            cell.alignment = _center()
            cell.border    = thin_border

    def _style_total_row(ws, row_num, n_cols):
        ws.row_dimensions[row_num].height = 16
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=True, color=C_TOTAL_FG, size=10)
            cell.fill      = _fill(C_TOTAL_BG)
            cell.alignment = _right() if c > 2 else _left()
            cell.border    = thin_border

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Workload Summary
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Workload Summary"
    ws1.freeze_panes = "C3"   # Freeze sub name + ID, keep header visible

    # Title banner
    ws1.merge_cells("A1:U1")
    title_cell = ws1["A1"]
    title_cell.value     = f"Cortex Cloud — Azure Workload Sizing  |  Generated: {generated_at}"
    title_cell.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=12)
    title_cell.fill      = _fill(C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 24

    # Column definitions: (header, field_key, width, number_format)
    wl_cols = [
        ("Subscription Name",  "sub_name",         32, "@"),
        ("Subscription ID",    "sub_id",            36, "@"),
        ("VMs Running",        "vms_running",        9, "#,##0"),
        ("AKS Nodes",          "aks_nodes",          9, "#,##0"),
        ("ARO Nodes",          "aro_nodes",          9, "#,##0"),
        ("ACI Containers",     "aci_containers",    10, "#,##0"),
        ("Container Apps",     "container_apps",    10, "#,##0"),
        ("Azure Functions",    "azure_functions",   11, "#,##0"),
        ("SQL Databases",      "sql_dbs",            9, "#,##0"),
        ("Cosmos DB",          "cosmos_dbs",         9, "#,##0"),
        ("Storage Accounts",              "storage_accounts",   12, "#,##0"),
        ("ACR Tags\n(Billable — All mode)", "acr_tag_count",     12, "#,##0"),
        ("ACR Manifests\n(Unique Digests)", "acr_manifest_count", 12, "#,##0"),
        ("ACR Images (Net)",               "acr_images_net",     12, "#,##0"),
        ("WL: VM/Nodes",       "wl_vm_nodes",       10, "#,##0"),
        ("WL: CaaS",           "wl_caas",            9, "#,##0"),
        ("WL: Functions",      "wl_functions",      10, "#,##0"),
        ("WL: Databases",      "wl_databases",      10, "#,##0"),
        ("WL: Storage",        "wl_storage",        10, "#,##0"),
        ("WL: Images",         "wl_images",          9, "#,##0"),
        ("TOTAL Workloads",    "wl_total",          13, "#,##0"),
    ]

    # Header row
    for ci, (hdr, _, width, _) in enumerate(wl_cols, 1):
        ws1.cell(row=2, column=ci).value = hdr
        _set_col_width(ws1, ci, width)
    _style_header_row(ws1, 2, len(wl_cols))

    # Data rows
    for ri, r in enumerate(rows, 3):
        ws1.row_dimensions[ri].height = 15
        for ci, (_, key, _, fmt) in enumerate(wl_cols, 1):
            cell = ws1.cell(row=ri, column=ci)
            cell.value        = r[key]
            cell.number_format = fmt
            cell.alignment    = _left() if ci <= 2 else _right()
            cell.border       = thin_border
            cell.font         = Font(name="Arial", size=10)
            # Alternating row shading
            if ri % 2 == 0:
                cell.fill = _fill(C_ALT_BG)

    # Totals row
    n_data = len(rows)
    tot_row = 3 + n_data
    ws1.cell(row=tot_row, column=1).value = "TOTAL"
    ws1.cell(row=tot_row, column=2).value = f"{n_data} subscription(s)"
    numeric_keys = [k for _, k, _, _ in wl_cols[2:]]
    for ci, (_, key, _, fmt) in enumerate(wl_cols[2:], 3):
        col_letter = get_column_letter(ci)
        cell = ws1.cell(row=tot_row, column=ci)
        cell.value         = f"=SUM({col_letter}3:{col_letter}{tot_row-1})"
        cell.number_format = fmt
    _style_total_row(ws1, tot_row, len(wl_cols))

    # Excel Table (for filtering)
    data_end_col = get_column_letter(len(wl_cols))
    tbl1 = Table(displayName="WorkloadSummary",
                 ref=f"A2:{data_end_col}{tot_row-1}")
    tbl1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws1.add_table(tbl1)

    # Conditional formatting — highlight high workload totals
    last_data_row = tot_row - 1
    total_col = get_column_letter(len(wl_cols))
    ws1.conditional_formatting.add(
        f"{total_col}3:{total_col}{last_data_row}",
        ColorScaleRule(
            start_type="min", start_color="E2EFDA",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max", end_color="FFC7CE"
        )
    )

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Log Ingestion
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Log Ingestion")
    ws2.freeze_panes = "C3"

    ws2.merge_cells("A1:N1")
    title2 = ws2["A1"]
    title2.value     = "Log Ingestion Sizing  |  Exact blob measurements only — no estimates"
    title2.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=12)
    title2.fill      = _fill(C_SECTION_BG)
    title2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 24

    log_cols = [
        ("Subscription Name",       "sub_name",          32, "@"),
        ("Subscription ID",         "sub_id",            36, "@"),
        ("EH Namespaces",           "eh_namespaces",      9, "#,##0"),
        ("EH w/ Capture",           "eh_with_capture",   10, "#,##0"),
        ("Flow Log Sources",        "flow_log_sources",  12, "#,##0"),
        ("Included GB/day",         "incl_gb_day",       13, "0.0000"),
        ("Audit GB/day",            "audit_gb_day",      13, "0.0000"),
        ("Flow GB/day",             "flow_gb_day",       12, "0.0000"),
        ("Runtime GB/day\n(Aud+Flw)","runtime_gb_day",   14, "0.0000"),
        ("Posture Add.\nGB/day",    "posture_add_gb",    13, "0.0000"),
        ("Runtime Add.\nGB/day",    "runtime_add_gb",    13, "0.0000"),
        ("Audit Source",            "audit_method",      40, "@"),
        ("Flow Source",             "flow_method",       40, "@"),
        ("Scan Status",             "scan_status",       10, "@"),
    ]

    for ci, (hdr, _, width, _) in enumerate(log_cols, 1):
        cell = ws2.cell(row=2, column=ci)
        cell.value     = hdr
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        _set_col_width(ws2, ci, width)
    _style_header_row(ws2, 2, len(log_cols), height=30)
    ws2.row_dimensions[2].height = 30

    for ri, r in enumerate(rows, 3):
        ws2.row_dimensions[ri].height = 15
        for ci, (_, key, _, fmt) in enumerate(log_cols, 1):
            cell = ws2.cell(row=ri, column=ci)
            val  = r[key]
            # Show "NOT AVAILABLE" string for None log values
            if val is None and key in ("audit_gb_day","flow_gb_day",
                                        "runtime_gb_day","posture_add_gb","runtime_add_gb"):
                cell.value         = "NOT AVAILABLE"
                cell.number_format = "@"
                cell.fill          = _fill(C_WARN_BG)
            else:
                cell.value         = val
                cell.number_format = fmt
                # Green background when within included allowance
                if key in ("posture_add_gb","runtime_add_gb") and val == 0.0:
                    cell.fill = _fill(C_GOOD_BG)
                elif key in ("posture_add_gb","runtime_add_gb") and val is not None and val > 0:
                    cell.fill = _fill(C_WARN_BG)
                elif ri % 2 == 0:
                    cell.fill = _fill(C_ALT_BG)
            cell.alignment = _left() if ci <= 2 else _right()
            cell.border    = thin_border
            cell.font      = Font(name="Arial", size=10)

    # Totals row for log sheet (only numeric cols that make sense to sum)
    tot2 = 3 + n_data
    ws2.cell(row=tot2, column=1).value = "TOTAL (measured subs only)"
    ws2.cell(row=tot2, column=2).value = f"{n_data} subscription(s)"
    sum_keys = {"eh_namespaces":3, "eh_with_capture":4, "flow_log_sources":5,
                "incl_gb_day":6, "audit_gb_day":7, "flow_gb_day":8}
    for key, ci in sum_keys.items():
        col_letter = get_column_letter(ci)
        cell = ws2.cell(row=tot2, column=ci)
        # Only sum numeric cells (skip NOT AVAILABLE strings)
        cell.value         = f"=SUMIF({col_letter}3:{col_letter}{tot2-1},\"<>NOT AVAILABLE\",{col_letter}3:{col_letter}{tot2-1})"
        cell.number_format = "0.0000" if "gb" in key else "#,##0"
    _style_total_row(ws2, tot2, len(log_cols))

    # Table for log sheet
    log_end_col = get_column_letter(len(log_cols))
    tbl2 = Table(displayName="LogIngestion",
                 ref=f"A2:{log_end_col}{tot2-1}")
    tbl2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws2.add_table(tbl2)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Grand Total Summary
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Grand Total")

    ws3.merge_cells("A1:D1")
    t3 = ws3["A1"]
    t3.value     = "Cortex Cloud — Grand Total Sizing Summary"
    t3.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=13)
    t3.fill      = _fill(C_HEADER_BG)
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 28
    _set_col_width(ws3, 1, 42)
    _set_col_width(ws3, 2, 18)
    _set_col_width(ws3, 3, 18)
    _set_col_width(ws3, 4, 40)

    def _gt_row(ws, row, label, value, note="", bold=False, bg=None, fmt="#,##0"):
        label_cell = ws.cell(row=row, column=1, value=label)
        val_cell   = ws.cell(row=row, column=2, value=value)
        note_cell  = ws.cell(row=row, column=4, value=note)
        for cell in (label_cell, val_cell, note_cell):
            cell.font      = Font(name="Arial", bold=bold, size=10,
                                  color=C_TOTAL_FG if bg == C_TOTAL_BG else "000000")
            cell.alignment = _left() if cell.column == 1 else _right()
            cell.border    = thin_border
            if bg:
                cell.fill = _fill(bg)
        val_cell.number_format = fmt
        ws.row_dimensions[row].height = 15

    def _gt_section(ws, row, title):
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=title)
        cell.font      = Font(name="Arial", bold=True, color=C_SECTION_FG, size=11)
        cell.fill      = _fill(C_SECTION_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        return row + 1

    r = 2
    r = _gt_section(ws3, r, "Subscription Workloads")
    _gt_row(ws3, r, "VMs / Nodes  (VMs + AKS + ARO)",       grand_totals["wl_vm_nodes"],   "1 workload per VM or node"); r+=1
    _gt_row(ws3, r, "CaaS  (ACI + Container Apps)",          grand_totals["wl_caas"],        "10 containers = 1 workload"); r+=1
    _gt_row(ws3, r, "Functions  (Azure Functions)",           grand_totals["wl_functions"],  "25 functions = 1 workload"); r+=1
    _gt_row(ws3, r, "Databases  (SQL + Cosmos DB)",           grand_totals["wl_databases"],  "2 databases = 1 workload"); r+=1
    _gt_row(ws3, r, "Storage  (Storage Accounts)",            grand_totals["wl_storage"],    "10 accounts = 1 workload"); r+=1
    _gt_row(ws3, r, "Images  (ACR net of free allowance)",    grand_totals["wl_images"],     "10 images = 1 workload"); r+=1
    _gt_row(ws3, r, "Subscription Total",                     grand_totals["wl_total"],
            "Sum of all subscription workloads", bold=True, bg=C_TOTAL_BG); r+=1

    saas_wl    = tenant.get("saas_user_workloads", 0)
    saas_total = tenant.get("total_saas_users", 0)
    r = _gt_section(ws3, r, "SaaS Users  (Entra ID — Tenant Level)")
    if tenant:
        _gt_row(ws3, r, "Entra ID Member users (enabled)", tenant.get("entra_member_users",0), "", fmt="#,##0"); r+=1
        _gt_row(ws3, r, "Entra ID Guest users (enabled)",  tenant.get("entra_guest_users",0),  "", fmt="#,##0"); r+=1
        _gt_row(ws3, r, "Total SaaS users",                saas_total, "Members + Guests",      fmt="#,##0"); r+=1
        _gt_row(ws3, r, "SaaS Workloads  (÷ 10)",          saas_wl,    "10 users = 1 workload"); r+=1
    else:
        ws3.cell(row=r, column=1).value = "Not scanned — run: python3 az-sizing.py --tenant-scan"
        ws3.cell(row=r, column=1).font  = Font(name="Arial", italic=True, color="888888", size=10)
        r+=1

    grand_total = grand_totals["wl_total"] + saas_wl
    grand_incl  = round(grand_total / 50, 4) if grand_total > 0 else 0.0
    r = _gt_section(ws3, r, "Grand Total")
    _gt_row(ws3, r, "GRAND TOTAL Workload Licenses",    grand_total,
            "Subscription + SaaS", bold=True, bg=C_TOTAL_BG); r+=1
    _gt_row(ws3, r, "Included log ingestion  (÷ 50)",   grand_incl,
            "GB/day included at no extra cost", fmt="0.0000"); r+=1

    r = _gt_section(ws3, r, "Log Ingestion Sizing")

    def _log_row(ws, row, label, value, note):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value if value is not None else "NOT AVAILABLE")
        nc = ws.cell(row=row, column=4, value=note)
        for cell in (lc, vc, nc):
            cell.font      = Font(name="Arial", size=10)
            cell.border    = thin_border
            cell.alignment = _left() if cell.column != 2 else _right()
        vc.number_format = "0.0000" if value is not None else "@"
        if value is None:
            vc.fill = _fill(C_WARN_BG)
        ws.row_dimensions[row].height = 15

    _log_row(ws3, r, "Audit Logs GB/day  (Posture SKU + Runtime SKU — same volume, not additive)",
             ingestion["tot_audit"] if not ingestion["miss_audit"] else None,
             "EH capture blobs or LAW proxy — 7-day avg" + (" ⚠ partial" if ingestion["miss_audit"] else "")); r+=1
    _log_row(ws3, r, "Flow Logs GB/day   (Runtime SKU only — not additive to Posture)",
             ingestion["tot_flow"] if not ingestion["miss_flow"] else None,
             "NSG/VNet flow-log blobs — prefix-based 7-day avg" + (" ⚠ partial" if ingestion["miss_flow"] else "")); r+=1
    _log_row(ws3, r, "Grand Total Included GB/day", grand_incl,
             "= Grand Total WL / 50"); r+=1

    # Posture — audit only
    if not ingestion["miss_audit"]:
        p_add = round(max(0.0, ingestion["tot_audit"] - grand_incl), 4)
        p_note = "✓ within included allowance" if p_add == 0 else f"{p_add:.4f} GB/day additional purchase needed"
    else:
        p_add  = None
        p_note = "Cannot determine — audit log volume not fully measured"
    _log_row(ws3, r, "Posture SKU additional GB/day  (Audit Logs only)", p_add, p_note); r+=1

    # Runtime — audit + flow (superset of Posture, not additive)
    if not ingestion["miss_audit"] and not ingestion["miss_flow"]:
        rt_total = round(ingestion["tot_audit"] + ingestion["tot_flow"], 4)
        r_add    = round(max(0.0, rt_total - grand_incl), 4)
        r_note   = "✓ within included allowance" if r_add == 0 else f"{r_add:.4f} GB/day additional"
    else:
        r_add  = None
        r_note = "Cannot determine — audit and/or flow not fully measured"
    _log_row(ws3, r, "Runtime SKU additional GB/day  (Audit + Flow — superset of Posture)", r_add, r_note); r+=1
    _log_row(ws3, r, "ⓘ Runtime includes Posture. These SKUs are not additive.", None,
             "A customer buying Runtime does not also need Posture separately."); r+=1

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 4 — Log Source Inventory  (one row per log source resource)
    # ════════════════════════════════════════════════════════════════════════
    # Build the inventory from two sources:
    #   a) Diagnostics records for FAILED sources (have resource name + container)
    #   b) Aggregate flow/EH/LAW summary for sub-level context on measured subs
    # This gives a row per individual resource rather than per subscription.

    ws4 = wb.create_sheet("Log Source Inventory")
    ws4.freeze_panes = "D3"

    n_inv_cols = 9
    ws4.merge_cells(f"A1:{get_column_letter(n_inv_cols)}1")
    t4 = ws4["A1"]
    t4.value     = "Log Source Inventory — one row per log source resource (flow logs, EH capture, LAW)"
    t4.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=12)
    t4.fill      = _fill(C_SECTION_BG)
    t4.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[1].height = 24

    inv_hdrs = [
        ("Subscription",      22),
        ("Source Type",       18),
        ("Resource Name",     28),
        ("Container / Path",  32),
        ("Status",            14),
        ("Category",          24),
        ("Impact",            26),
        ("Fix",               72),
        ("Sub ID",            36),
    ]
    for ci, (hdr, width) in enumerate(inv_hdrs, 1):
        c = ws4.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _left()
        c.border    = thin_border
        _set_col_width(ws4, ci, width)
    ws4.row_dimensions[2].height = 18

    _IMPACT_LABELS = {
        "flow_log_unmeasured":        "Flow log volume not measured",
        "audit_log_unmeasured":       "Audit log volume not measured",
        "law_volume_unmeasured":      "LAW workspace volume not measured",
        "acr_image_count_unmeasured": "ACR image count not collected",
        "eh_capture_unmeasured":      "EH capture volume not measured",
    }
    _SOURCE_LABELS = {
        "storage_account":    "VNet/NSG Flow Log",
        "event_hub_capture":  "EH Capture (Audit)",
        "law":                "Log Analytics WS",
        "acr":                "Container Registry",
        "event_hub":          "Event Hub",
    }
    _CAT_SHORT = {
        "rbac_missing_sbdr":       "SBDR missing",
        "rbac_missing_law_reader": "LAW Reader missing",
        "rbac_missing_eh_reader":  "EH Reader missing",
        "rbac_missing_acr_pull":   "AcrPull missing",
        "firewall_blocked":        "Firewall blocked",
        "private_endpoint_only":   "Private endpoint",
        "ssl_intercept":           "SSL intercept",
        "tenant_mismatch":         "Tenant mismatch",
        "container_not_present":   "Container absent",
        "not_configured":          "Not configured",
    }

    inv_rows = []

    # Failed sources from structured diagnostics
    for r in rows:
        for d in r.get("diagnostics", []):
            inv_rows.append({
                "sub":        r["sub_name"],
                "src_type":   _SOURCE_LABELS.get(d.get("resource_type", ""), d.get("resource_type", "")),
                "rname":      d.get("resource_name", ""),
                "subpath":    d.get("resource_sub_path", ""),
                "status":     "❌ Not measured",
                "category":   _CAT_SHORT.get(d.get("issue_category", ""), d.get("issue_category", "")),
                "impact":     _IMPACT_LABELS.get(d.get("impact", ""), d.get("impact", "")),
                "fix":        d.get("fix", ""),
                "sub_id":     d.get("subscription_id", r["sub_id"]),
            })

    # Measured flow log sources — per flow log resource (from _flow_log_inventory)
    for r in rows:
        for entry in r.get("flow_log_inventory", []):
            target_name = entry.get("target_id", "").rsplit("/", 1)[-1] or "(unknown target)"
            fl_label    = entry.get("flow_log_name", "")
            shared_note = "  (shared acct)" if entry.get("shared_account") else ""
            inv_rows.append({
                "sub":      r["sub_name"],
                "src_type": "VNet/NSG Flow Log",
                "rname":    fl_label,
                "subpath":  f"→ {target_name}  |  {entry.get('container', '')}",
                "status":   f"✅ Measured  {entry.get('gb_day', 0.0):.4f} GB/day{shared_note}",
                "category": "",
                "impact":   "",
                "fix":      "",
                "sub_id":   r["sub_id"],
            })

    # Measured EH capture paths — per namespace/hub (from _eh_capture_inventory)
    for r in rows:
        for entry in r.get("eh_capture_inventory", []):
            ns   = entry.get("ns_name", "(unknown namespace)")
            hub  = entry.get("hub_name", "(unknown hub)")
            inv_rows.append({
                "sub":      r["sub_name"],
                "src_type": "EH Capture (Audit)",
                "rname":    f"{ns} / {hub}",
                "subpath":  f"{entry.get('storage_account', '')}  |  {entry.get('container', '')}",
                "status":   f"✅ Measured  {entry.get('gb_day', 0.0):.4f} GB/day",
                "category": "",
                "impact":   "",
                "fix":      "",
                "sub_id":   r["sub_id"],
            })

    if not inv_rows:
        ws4.cell(row=3, column=1).value = "No log sources discovered or all sources measured without issues."
        ws4.cell(row=3, column=1).font  = Font(name="Arial", italic=True, color="228B22", size=10)
    else:
        STATUS_BG = {"❌": C_WARN_BG, "✅": "CCFFCC", "⚠": "FFEB9C"}
        for ri, iv in enumerate(inv_rows, 3):
            vals = [iv["sub"], iv["src_type"], iv["rname"], iv["subpath"],
                    iv["status"], iv["category"], iv["impact"], iv["fix"], iv["sub_id"]]
            stat_bg = next((v for k, v in STATUS_BG.items() if iv["status"].startswith(k)), C_ALT_BG if ri % 2 == 0 else "FFFFFF")
            ws4.row_dimensions[ri].height = 30
            for ci, val in enumerate(vals, 1):
                cell = ws4.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(ci in (4, 8)))
                if ci == 5:
                    cell.fill = _fill(stat_bg)
                elif ri % 2 == 0:
                    cell.fill = _fill(C_ALT_BG)

        inv_end = get_column_letter(len(inv_hdrs))
        tbl4 = Table(displayName="LogSourceInventory",
                     ref=f"A2:{inv_end}{2+len(inv_rows)}")
        tbl4.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws4.add_table(tbl4)

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 5 — Diagnostics  (one row per structured DiagnosticRecord)
    # ════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Diagnostics")
    ws5.freeze_panes = "D3"

    # Collect ALL structured diagnostic records from every subscription
    all_diags = []
    diag_id   = 1
    for r in rows:
        for d in r.get("diagnostics", []):
            all_diags.append({**d, "_id": f"DIAG-{diag_id:03d}"})
            diag_id += 1

    n_diag_cols = 9
    ws5.merge_cells(f"A1:{get_column_letter(n_diag_cols)}1")
    t5 = ws5["A1"]
    t5.value     = (
        f"Diagnostics — {len(all_diags)} issue(s) across {n_data} subscription(s)  |  "
        "Sort or filter any column to triage by category, subscription, or resource"
    )
    t5.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=12)
    t5.fill      = _fill(C_HEADER_BG)
    t5.alignment = Alignment(horizontal="left", vertical="center")
    ws5.row_dimensions[1].height = 24

    diag_hdrs = [
        ("ID",               8),
        ("Subscription",    26),
        ("Resource Type",   20),
        ("Resource Name",   28),
        ("Container/Path",  28),
        ("Category",        26),
        ("Impact",          30),
        ("Fix",             80),
        ("Sub ID",          36),
    ]
    for ci, (hdr, width) in enumerate(diag_hdrs, 1):
        c = ws5.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _left()
        c.border    = thin_border
        _set_col_width(ws5, ci, width)
    ws5.row_dimensions[2].height = 18

    _CAT_COLORS = {
        "rbac_missing_sbdr":       "FFC7CE",   # red
        "rbac_missing_law_reader": "FFC7CE",
        "rbac_missing_eh_reader":  "FFC7CE",
        "rbac_missing_acr_pull":   "FFC7CE",
        "firewall_blocked":        "FFEB9C",   # amber
        "private_endpoint_only":   "DDEBF7",   # blue
        "ssl_intercept":           "FFEB9C",
        "tenant_mismatch":         "FFC7CE",
        "container_not_present":   "E2EFDA",   # green (informational)
        "not_configured":          "DDEBF7",
    }

    if not all_diags:
        ws5.cell(row=3, column=1).value = "✅  No issues detected — all log sources measured successfully."
        ws5.cell(row=3, column=1).font  = Font(name="Arial", italic=True, color="228B22", size=10)
    else:
        for ri, d in enumerate(all_diags, 3):
            cat     = d.get("issue_category", "")
            cat_bg  = _CAT_COLORS.get(cat, "FFFFFF")
            cat_lbl = _CAT_SHORT.get(cat, cat)
            imp_lbl = _IMPACT_LABELS.get(d.get("impact", ""), d.get("impact", ""))
            vals = [
                d["_id"],
                d.get("subscription_name", ""),
                _SOURCE_LABELS.get(d.get("resource_type", ""), d.get("resource_type", "")),
                d.get("resource_name", ""),
                d.get("resource_sub_path", ""),
                cat_lbl,
                imp_lbl,
                d.get("fix", ""),
                d.get("subscription_id", ""),
            ]
            ws5.row_dimensions[ri].height = 36
            for ci, val in enumerate(vals, 1):
                cell = ws5.cell(row=ri, column=ci, value=val)
                cell.font      = Font(name="Arial", size=10,
                                      bold=(ci == 6))   # bold category
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=(ci in (5, 8)))
                if ci == 6:
                    cell.fill = _fill(cat_bg)
                elif ri % 2 == 0:
                    cell.fill = _fill(C_ALT_BG)

        diag_end = get_column_letter(len(diag_hdrs))
        tbl5 = Table(displayName="Diagnostics",
                     ref=f"A2:{diag_end}{2+len(all_diags)}")
        tbl5.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium3", showFirstColumn=False,
            showLastColumn=False, showRowStripes=False, showColumnStripes=False
        )
        ws5.add_table(tbl5)

    # ── Final save ───────────────────────────────────────────────────────────
    wb.save(out_path)
    print(f"\n  ✅  Excel workbook saved: {out_path}")
    print(f"     Sheets: Workload Summary | Log Ingestion | Grand Total | Log Source Inventory | Diagnostics")


# ──────────────────────────────────────────────────────────────────────────────
# Orchestration
# ──────────────────────────────────────────────────────────────────────────────
def print_summary(results: dict, state: dict, tenant: dict,
                  failed_only: bool, xlsx_path: str | None) -> None:
    if not results:
        print("  No results to display.  Run az-sizing.py --resume first.")
        return

    if state:
        sc = {}
        for r_row in state.values():
            s = r_row.get("status", "unknown")
            sc[s] = sc.get(s, 0) + 1
        print("  Scan state: " + "  ".join(f"{s}={n}" for s, n in sorted(sc.items())))

    # Sort alphabetically by name, empty subs last
    def _sort_key(item):
        raw = item[1].get("raw_counts", {})
        total_resources = sum([
            raw.get("vm_running", 0), raw.get("aks_nodes", 0),
            raw.get("storage_accounts", 0), raw.get("azure_functions", 0),
        ])
        name = item[1].get("name", "").lower()
        # Push zero-resource subs to bottom
        return (0 if total_resources > 0 else 1, name)

    sorted_items = sorted(results.items(), key=_sort_key)

    if failed_only:
        sorted_items = [
            (sid, p) for sid, p in sorted_items
            if state.get(sid, {}).get("status") == "failed"
        ]

    rows = build_rows(sorted_items, state)

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    grand_totals = print_section1(rows)
    ingestion    = print_section2(rows)
    print_section3(grand_totals, ingestion, tenant)

    if xlsx_path:
        export_excel(rows, tenant, ingestion, grand_totals, xlsx_path, generated_at)

    # Failed subs footer
    if state and not failed_only:
        missing_failed = [
            r_row for sid, r_row in state.items()
            if r_row.get("status") == "failed" and sid not in results
        ]
        if missing_failed:
            print(f"\n{THIN}")
            print(f"  {len(missing_failed)} failed subscription(s) with no results data:")
            for r_row in missing_failed:
                print(f"    {r_row['name']:<42} {r_row['sub_id']}  "
                      f"error: {r_row.get('last_error', 'n/a')}")
            print(f"  Retry: python3 az-sizing.py --resume --retry-failed")
            print(THIN)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        prog="az-summary.py",
        description="Cortex Cloud Azure sizing — consolidated summary + Excel export",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 az-summary.py
  python3 az-summary.py --results azure_results.json --tenant azure_tenant.json
  python3 az-summary.py --results azure_results.json --xlsx sizing.xlsx
  python3 az-summary.py --results run1.json run2.json --state azure_state.jsonl
  python3 az-summary.py --failed-only --state azure_state.jsonl
""",
    )
    parser.add_argument(
        "--results", nargs="*", default=["azure_results.json"], metavar="FILE",
        help="azure_results.json file(s)  [default: azure_results.json]",
    )
    parser.add_argument(
        "--state", default=None, metavar="FILE",
        help="azure_state.jsonl — show subscription scan status alongside results",
    )
    parser.add_argument(
        "--tenant", default=None, metavar="FILE",
        help="azure_tenant.json from az-sizing.py --tenant-scan  (SaaS Users)",
    )
    parser.add_argument(
        "--xlsx", default=None, metavar="FILE",
        help="Export results to Excel workbook  (e.g. sizing_output.xlsx)",
    )
    parser.add_argument(
        "--failed-only", action="store_true",
        help="Show only failed subscriptions  (requires --state)",
    )
    args = parser.parse_args()

    print(f"\n{SEP}")
    print("  CORTEX CLOUD  —  AZURE WORKLOAD SIZING SUMMARY")
    print(f"  Generated : {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC")
    print(SEP)
    print("\nLoading data ...")

    results = load_results(args.results)
    state   = load_state(args.state)
    tenant  = load_tenant(args.tenant)

    if state:
        print(f"  Loaded state for {len(state)} subscription(s) from {args.state}")
    if tenant:
        print(f"  Loaded tenant data  (tenant_id={tenant.get('tenant_id','unknown')}, "
              f"scanned={tenant.get('timestamp','—')})")

    if args.failed_only and not state:
        print("  --failed-only requires --state <azure_state.jsonl>")
        return

    print_summary(results, state or {}, tenant, args.failed_only, args.xlsx)


if __name__ == "__main__":
    main()
